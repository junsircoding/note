"""
使用 openpyxl 对 Excel 进行读写
参考文章：https://blog.csdn.net/sinat_28576553/article/details/81275650
"""

import openpyxl

# 加载Excel
workbook = openpyxl.load_workbook('学生成绩单.xlsx')
# 查看Sheet名称
print(workbook.sheetnames)
# 获取某个Sheet
sheet = workbook['Sheet1']
# 获取某一单元格数据
print(sheet['A1'].value)
print(sheet.cell(2, 2).value)
# 修改数据
sheet.cell(2, 3, '0')
sheet["B2"] = "Peking"
# 保存修改
workbook.save(filename='DataSource\\myfile.xlsx')
