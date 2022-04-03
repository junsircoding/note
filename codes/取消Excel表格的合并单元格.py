"""
`openpyxl` 取消 `Excel` 表格的**合并单元格**，并自动填充数据
"""

from openpyxl import load_workbook

# 1. 加载数据
work_book = load_workbook(filename="data2.xlsx", read_only=False)
sheet = work_book["Sheet1"]

# 2. 找出所有的合并单元格的索引信息
mc_range_list = [str(item) for item in sheet.merged_cells.ranges]

# 3. 批量取消合并单元格，填充数据
for mc_range in mc_range_list:
    # 取得左上角值的坐标
    top_left, bot_right = mc_range.split(":") # ["A1", "A12"]
    top_left_col, top_left_row = sheet[top_left].column, sheet[top_left].row # (1, 1,)
    bot_right_col, bot_right_row = sheet[bot_right].column, sheet[bot_right].row # (1, 12,)
    # 记下该合并单元格的值
    cell_value = sheet[top_left].value # 忍者
    # 取消合并单元格
    sheet.unmerge_cells(mc_range)
    # 批量给子单元格赋值
    # 遍历列
    for col_idx in range(top_left_col, bot_right_col+1):
        # 遍历行
        for row_idx in range(top_left_row+1, bot_right_row+1):
            sheet[f"{chr(col_idx+64)}{row_idx}"] = cell_value

# 4. 保存更改            
work_book.save("data3.xlsx")
