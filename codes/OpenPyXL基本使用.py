"""
OpenPyXL 基本使用
"""

import openpyxl

# TODO: 创建新表格
# 生成一个新的工作簿
new_work_book = openpyxl.Workbook()
# 得到当前激活的表格
new_sheet = new_work_book.active
# 创建一个新表格
new_work_book.create_sheet("new_sheet_1")
new_sheet_1 = new_work_book["new_sheet_1"]
# 给新表格命名
new_sheet.title = "Sheet1"
# 直接根据行列坐标赋值
new_sheet["F4"] = "数据"
# 将 ASCII 数字转为对应大写字母, 从 0 开始 `chr(n+64)`
# 批量追加数据
data_list = [
    ["春", "眠", "不", "觉", "晓"],
    ["处", "处", "闻", "啼", "鸟"],
    ["夜", "来", "风", "雨", "声"],
    ["花", "落", "知", "多", "少"]
]
for new_row in data_list:
    new_sheet.append(new_row)
new_work_book.save("./data2.xlsx")


# TODO: 读取现有表格
# 读取工作簿对象
work_book = openpyxl.load_workbook("./data.xlsx")
# 获取表格名称列表
sheet_names = work_book.sheetnames
# 拿到某个表格
sheet = work_book["数据源一"]
# 表格名称
print(sheet.title)
# 表格最大行数
print(sheet.max_row)
# 表格最大列数
print(sheet.max_column)
# 遍历表格数据
for row in sheet.rows:
    row_col_nums = []
    for cell in row:
        row_num = cell.row  # 行号
        col_num = cell.column  # 列号
        cell_value = cell.value  # 单元格数据
        row_col_nums.append(f"[{col_num}-{row_num}]:{cell_value}")
    print(row_col_nums)
# 获取某一单元格数据
print(sheet['A1'].value)
print(sheet.cell(2, 2).value)
# 修改数据
sheet.cell(2, 3, '0')
sheet["B2"] = "Peking"
# 保存修改
work_book.save(filename='DataSource\\myfile.xlsx')
